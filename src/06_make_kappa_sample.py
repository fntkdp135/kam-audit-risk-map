# -*- coding: utf-8 -*-
"""⑥ 사람 검증용 무작위 표본 생성 (blind).

자동 분류(규칙+LLM) 결과를 감춘 상태로 60건을 뽑아 사람이 직접 분류하게 한다.
정답표(기계 분류)는 별도 파일로 분리 보관하며, 사람 분류가 끝난 뒤에만 대조한다.
표본은 층화하지 않은 단순 무작위 추출(seed 고정) — 감사의 표본검사와 같은 방식.
"""
import os, sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "processed")
RES = os.path.join(ROOT, "data", "results")
os.makedirs(RES, exist_ok=True)

N = 60
SEED = 20260731

TYPES = ["수익인식", "영업권 손상", "콘텐츠·제작비 자산", "종속·관계기업투자 평가",
         "유형·비유동자산 손상", "금융상품 공정가치·평가", "매출채권·대손", "재고자산",
         "리스", "특수관계자거래", "계속기업·유동성", "이연법인세", "사업결합",
         "충당부채·소송·우발", "판단 불가"]

DEFS = [
    ("수익인식", "매출의 실재성·기간귀속·정확성, 변동대가, 진행률/투입법, 계약부채·이연수익 등 수익 관련 쟁점"),
    ("영업권 손상", "영업권 또는 현금창출단위(CGU) 손상평가"),
    ("콘텐츠·제작비 자산", "판권·콘텐츠·개발비 등 제작 관련 자산의 손상·자산화. "
                    "계정과목이 무형자산이든 선급금·미니멈개런티든 회계쟁점이 같으면 여기"),
    ("종속·관계기업투자 평가", "종속기업·관계기업 투자주식의 손상·평가·처분, 지배력 상실 회계처리"),
    ("유형·비유동자산 손상", "유형자산·투자부동산·사용권자산 등의 손상·실재성·평가"),
    ("금융상품 공정가치·평가", "파생상품, 전환사채·상환전환우선주 등 복합금융상품, 수준3 공정가치"),
    ("매출채권·대손", "매출채권 회수가능성, 대손충당금, 기대신용손실"),
    ("재고자산", "재고자산 평가·진부화"),
    ("리스", "리스 회계처리·리스계약의 완전성"),
    ("특수관계자거래", "특수관계자 거래의 발생사실·표시·공시, 내부거래, 자금대여"),
    ("계속기업·유동성", "계속기업 가정, 유동성 위험, 자금조달"),
    ("이연법인세", "이연법인세자산의 실현가능성"),
    ("사업결합", "사업결합·합병·물적분할·거래가격배분(PPA)·중단영업"),
    ("충당부채·소송·우발", "충당부채, 우발부채, 소송, 손해배상, 반품·환불"),
    ("판단 불가", "항목명이 잘렸거나 내용이 부족해 유형을 정할 수 없는 경우"),
]


def main():
    k = pd.read_csv(os.path.join(OUT, "kam_typed.csv"), dtype={"rcept_no": str},
                    encoding="utf-8-sig")
    # 연결/별도 중복 KAM은 한쪽만 (같은 항목을 두 번 분류하게 하지 않기 위함)
    k = k.drop_duplicates(subset=["corp_code", "fy", "kam_title"], keep="first")
    s = k.sample(n=min(N, len(k)), random_state=SEED).reset_index(drop=True)
    s.insert(0, "번호", range(1, len(s) + 1))

    # 정답표(기계 분류)는 분리 보관
    s[["번호", "corp_code", "fy", "기업명", "kam_title", "kam_type", "type_source"]] \
        .to_csv(os.path.join(OUT, "kappa_machine_key.csv"), index=False, encoding="utf-8-sig")

    wb = Workbook()
    ws = wb.active
    ws.title = "분류"
    heads = ["번호", "기업명", "사업연도", "KAM 항목명", "선정 사유(발췌)", "감사절차(발췌)", "분류"]
    ws.append(heads)
    for c in range(1, len(heads) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F4F6F")
        cell.alignment = Alignment(vertical="center")
    for r in s.itertuples():
        ws.append([r.번호, r.기업명, int(r.fy), str(r.kam_title),
                   str(r.reason_text)[:700] if pd.notna(r.reason_text) else "",
                   str(r.procedure_text)[:400] if pd.notna(r.procedure_text) else "", ""])

    widths = [6, 18, 9, 34, 78, 52, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row[0].row].height = 96

    dv = DataValidation(type="list", formula1='"%s"' % ",".join(TYPES), allow_blank=True)
    dv.error = "목록에서 선택해 주세요"
    ws.add_data_validation(dv)
    dv.add(f"G2:G{ws.max_row}")
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("유형 정의")
    ws2.append(["유형", "정의"])
    for c in range(1, 3):
        ws2.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws2.cell(1, c).fill = PatternFill("solid", fgColor="2F4F6F")
    for t, d in DEFS:
        ws2.append([t, d])
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 92
    for row in ws2.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[row[0].row].height = 34

    path = os.path.join(RES, "KAM분류_사람검증_60건.xlsx")
    wb.save(path)
    print(f"저장: {path}")
    print(f"표본 {len(s)}건 (seed={SEED}, 단순 무작위)")
    print(f"정답표: {os.path.join(OUT,'kappa_machine_key.csv')} (대조 전까지 열지 말 것)")
    print("\n표본의 기계 분류 분포(참고용, 사람 분류 후 대조):")
    print(s.kam_type.value_counts().to_string())


if __name__ == "__main__":
    main()
