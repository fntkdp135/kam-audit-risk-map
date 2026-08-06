# -*- coding: utf-8 -*-
"""⑫ FY2025 추가 표본 생성 (κ 검증 범위 보강).

원표본 60건은 FY2019~2024 모집단에서 뽑은 것이라 FY2025 항목은 뽑힐 기회가 없었다.
검증 범위를 모집단 전체로 넓히기 위해 FY2025에서만 추가 표본을 뽑는다.
형식·유형 정의는 06과 동일하게 유지해 사람이 같은 기준으로 분류하도록 한다.
"""
import os, sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "data", "processed")
RES = os.path.join(ROOT, "data", "results")

TARGET_FY = 2025
N = 15
SEED = 20260807

TYPES = ["수익인식", "영업권 손상", "콘텐츠·제작비 자산", "종속·관계기업투자 평가",
         "유형·비유동자산 손상", "금융상품 공정가치·평가", "매출채권·대손", "재고자산",
         "리스", "특수관계자거래", "계속기업·유동성", "이연법인세", "사업결합",
         "충당부채·소송·우발", "판단 불가"]

DEFS = [
    ("수익인식", "매출의 실재성·기간귀속·정확성, 변동대가, 진행률/투입법, 계약부채·이연수익 등 수익 관련 쟁점"),
    ("영업권 손상", "영업권 또는 현금창출단위(CGU) 손상평가"),
    ("콘텐츠·제작비 자산", "판권·콘텐츠·개발비 등 제작 관련 자산의 손상·자산화. "
                    "계정과목이 무형자산이든 선급금·미니멈개런티든 회계쟁점이 같으면 여기"),
    ("종속·관계기업투자 평가", "종속기업·관계기업 투자주식의 손상·평가·처분, 지배력 상실 회계처리"),
    ("유형·비유동자산 손상", "유형자산·투자부동산·사용권자산 등의 손상·실재성·평가"),
    ("금융상품 공정가치·평가", "파생상품, 전환사채·상환전환우선주 등 복합금융상품, 수준3 공정가치"),
    ("매출채권·대손", "매출채권 회수가능성, 대손충당금, 기대신용손실"),
    ("재고자산", "재고자산 평가·진부화"),
    ("리스", "리스 회계처리·리스계약의 완전성"),
    ("특수관계자거래", "특수관계자 거래의 발생사실·표시·공시, 내부거래, 자금대여"),
    ("계속기업·유동성", "계속기업 가정, 유동성 위험, 자금조달"),
    ("이연법인세", "이연법인세자산의 실현가능성"),
    ("사업결합", "사업결합·합병·물적분할·거래가격배분(PPA)·중단영업"),
    ("충당부채·소송·우발", "충당부채, 우발부채, 소송, 손해배상, 반품·환불"),
    ("판단 불가", "항목명이 잘렸거나 내용이 부족해 유형을 정할 수 없는 경우"),
]


def main():
    k = pd.read_csv(os.path.join(OUT, "kam_typed.csv"), dtype={"corp_code": str},
                    encoding="utf-8-sig")
    k["corp_code"] = k.corp_code.str.zfill(8)
    k = k.drop_duplicates(subset=["corp_code", "fy", "kam_title"], keep="first")
    k = k[(k.fy == TARGET_FY) & (k.kam_type != "미분류")]

    s = k.sample(n=min(N, len(k)), random_state=SEED).reset_index(drop=True)
    s.insert(0, "번호", range(1, len(s) + 1))

    s[["번호", "corp_code", "fy", "기업명", "kam_title", "kam_type", "type_source"]] \
        .to_csv(os.path.join(OUT, "kappa_machine_key_fy2025.csv"),
                index=False, encoding="utf-8-sig")

    wb = Workbook()
    ws = wb.active
    ws.title = "분류"
    heads = ["번호", "기업명", "사업연도", "KAM 항목명", "선정 사유(발췌)", "감사절차(발췌)", "분류"]
    ws.append(heads)
    for c in range(1, len(heads) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F4F6F")
        cell.alignment = Alignment(vertical="center")
    for r in s.itertuples():
        ws.append([r.번호, r.기업명, int(r.fy), str(r.kam_title),
                   str(r.reason_text)[:700] if pd.notna(r.reason_text) else "",
                   str(r.procedure_text)[:400] if pd.notna(r.procedure_text) else "", ""])

    for i, w in enumerate([6, 18, 9, 34, 78, 52, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row[0].row].height = 96

    dv = DataValidation(type="list", formula1='"%s"' % ",".join(TYPES), allow_blank=True)
    dv.error = "목록에서 선택해 주세요"
    ws.add_data_validation(dv)
    dv.add(f"G2:G{ws.max_row}")
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("유형 정의")
    ws2.append(["유형", "정의"])
    for c in range(1, 3):
        ws2.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws2.cell(1, c).fill = PatternFill("solid", fgColor="2F4F6F")
    for t, d in DEFS:
        ws2.append([t, d])
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 92
    for row in ws2.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[row[0].row].height = 34

    path = os.path.join(RES, f"KAM분류_사람검증_FY{TARGET_FY}_{len(s)}건.xlsx")
    wb.save(path)
    print(f"저장: {path}")
    print(f"FY{TARGET_FY} 모집단 {len(k)}개 중 {len(s)}건 추출 (seed={SEED}, 단순 무작위)")
    print("\n표본의 기계 분류 분포(대조 전까지 열지 말 것):")
    print(s.kam_type.value_counts().to_string())


if __name__ == "__main__":
    main()
